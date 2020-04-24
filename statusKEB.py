# -*- coding: utf-8 -*-
# Заполняем поле статуса в монго конвертируя его из Excel

import sys, argparse
from _datetime import datetime, timedelta, date
import time
import os
from mysql.connector import MySQLConnection, Error
from collections import OrderedDict
import openpyxl
from pymongo import MongoClient
import psycopg2

from lib import read_config, s, l

EVA_STATUS = {
'STATUS_NONE' : 0, # Utils DEFAULT_VALUE
'STATUS_NEW' : 20, # Новая заявка
'STATUS_QUEUED' : 100, # Заявка отправлена в очередь
'STATUS_CONFIRM' : 110, # Введен СМС код
'STATUS_RETRY' : 120, # Запрошена повторная СМС
'STATUS_PROCESSING' : 130, # В процессе
'STATUS_APPROVED' : 140, # Одобрена
'STATUS_PRE_APPROVED' : 150, # Предварительно одобрена
'STATUS_DONE' : 200, # Завершено успешно
'STATUS_DELETED' : 400, # Удалена
'STATUS_UNKNOWN' : 410, # Неизвестный статус
'STATUS_DENIED' : 430, # Отказ
'EVENT_UPDATE' : 10, # Анкета отредактирована
'STATUS_DEBUG' : 500, # Отладка
'STATUS_DRAFT' : 510, # Отложена
'STATUS_TRANSACTION_ERROR' : 420, # Ошибка выгрузки
'STATUS_HAS_ERROR' : 470, # Ошибка в заявке
'REMOTE_STATUS_AWAITING' : 600, # Ожидает оплаты
'REMOTE_STATUS_PAYED' : 610, # Оплачено
'REMOTE_STATUS_DONE' : 620, # Услуга получена
'STATUS_ISSUED' : 210, # Займ выдан
'STATUS_DOUBLE_ISSUED' : 220, # Займ выдан повторно
'STATUS_ISSUED_CALLCENTER' : 230, # Займ выдан через call-центр
'STATUS_COMPLETED' : 160, # Заявка заполнена
'STATUS_SEND_ANKETA' : 170, # Анкета успешно отправлена
'STATUS_FILE_ERROR' : 180, # Ошибка отправки файлов
'STATUS_SEND_FILE' : 190, # Файлы успешно отправлены
'STATUS_ERROR' : 50, # Ошибка
'RUSTELECOM_STATUS_PROCESSING' : 310, # Заявка отправлена
'RUSTELECOM_STATUS_SEND_SCANS' : 320, # Сканы отправлены
'RUSTELECOM_STATUS_SEND_DOCUMENT' : 330, # Документы отправлены
'RUSTELECOM_STATUS_ERROR' : 340, # Ошибка
'RUSTELECOM_STATUS_SEND_SMS' : 350, # Смс отправлена
'RUSTELECOM_STATUS_DONE' : 360, # Завершено
'RUSTELECOM_STATUS_SEND_CODE' : 370, # ЕСИА код отправлен
'STATUS_CLIENT_DENIAL' : 440, # Отказ клиента
'STATUS_CLOSED' : 450, # Закрыта
'STATUS_EXPIRED' : 460, # Истек срок действия решения Банка
'STATUS_APP_INSTALLED' : 240, # Приложение установлено
'STATUS_ACCOUNT_REPLENISHED' : 250, # Счет пополнен
'STATUS_ACTIVATED' : 260, # Карта активирована
'STATUS_ALFABANK_100_CREATED' : 1100, # Заявка создана
'STATUS_ALFABANK_100_PRESCORING_VALID' : 1200, # Пройден прескоринг
'STATUS_ALFABANK_100_PRESCORING_FAILED' : 1210, # Не пройден прескоринг
'STATUS_ALFABANK_100_SCORING_VALID' : 1300, # Пройден скоринг
'STATUS_ALFABANK_100_SCORING_FAILED' : 1310, # Не пройден скоринг
'STATUS_ALFABANK_100_DONE' : 1500, # Карта выдана
'STATUS_ALFABANK_100_ACTIVATED' : 1600, # Карта активирована
'ROCKETBANK_DEBIT_CARD_REFERRAL_NONE' : 700, # Карта не выдана
'ROCKETBANK_DEBIT_CARD_REFERRAL_DONE' : 710, # Карта выдана
'ROCKETBANK_VIRTUAL_DEBIT_CARD_REFERRAL_DONE' : 750, # Карта выдана
'ROCKETBANK_VIRTUAL_DEBIT_CARD_REFERRAL_ACTIVATE' : 760, # Карта активирована
'CREDITEUROPEBANK_CREDIT_CARD_REFERRAL_ANKETA_FILLED' : 800, # Анкета успешно заполнена
'ROSBANK_REFERRAL_APPROVED' : 900, # Одобрено
'ROSBANK_REFERRAL_DENIED' : 910, # Отклонено
'ROSBANK_REFERRAL_AWAITING' : 920, # В ожидании
'ROSBANK_REFERRAL_SEND' : 990, # Заявка передана
'OPENBANK_REFERRAL_CONFIRM' : 2000, # Принято
'OPENBANK_REFERRAL_PROCESSING' : 2100, # В обработке
'OPENBANK_REFERRAL_DENIED' : 2200, # Отклонено
'UBRR_API_PROCESSING' : 850, # В обработке
'UBRR_API_ERROR_DOUBLE' : 860, # Дубль заявки
'UBRR_API_ERROR_VALIDATE' : 870, # Данные не валидны
'UBRR_API_CLAIM_SEND' : 880, # Заявка передана
}

COLUMNS = ['UTM_CAMPAIGN','RESULT', 'PRESCORE', 'LIMIT', 'CARD_STATUS']

STATUSES = {}

STATUSES['RESULT'] = {
'отклонено скорингом': EVA_STATUS['STATUS_DENIED'],
'отклонена банком': EVA_STATUS['STATUS_DENIED'],
'не соответствует требованиям': EVA_STATUS['STATUS_DENIED'],
'продукт не нужен': EVA_STATUS['STATUS_DENIED'],
'отказ клиента': EVA_STATUS['STATUS_DENIED'],
'регион без представительства КЕБ': EVA_STATUS['STATUS_DENIED'],
'дубликат (предыдущая заявка)': EVA_STATUS['STATUS_DENIED'],
'карта выдана': EVA_STATUS['STATUS_ISSUED'],
'активирована карта': EVA_STATUS['STATUS_ACTIVATED'],
}

STATUSES['LIMIT'] = {
'отклонено скорингом': EVA_STATUS['STATUS_DENIED'],
'отклонена банком': EVA_STATUS['STATUS_DENIED'],
'не соответствует требованиям': EVA_STATUS['STATUS_DENIED'],
'продукт не нужен': EVA_STATUS['STATUS_DENIED'],
'отказ клиента': EVA_STATUS['STATUS_DENIED'],
'регион без представительства КЕБ': EVA_STATUS['STATUS_DENIED'],
'дубликат (предыдущая заявка)': EVA_STATUS['STATUS_DENIED'],
'карта выдана': EVA_STATUS['STATUS_ISSUED'],
'активирована карта': EVA_STATUS['STATUS_ACTIVATED'],
}

STATUSES['PRESCORE'] = {
'отказ': EVA_STATUS['STATUS_DENIED'],
'одобрено': EVA_STATUS['STATUS_APPROVED'],
}

# !!! Обрабатывается в последнюю очередь, ДОБАВЛЯТЬ только исключающие обработку статусы
STATUSES['CARD_STATUS'] = {
    'карта закрыта': True
}

def filter_x00(inp):
    inp = s(inp)
    inp = inp.replace('_x0020_',' ')
    inp = inp.replace('_X0020_',' ')
    while inp.upper().find('_X0') > -1:
        if inp.find('_x0') > -1:
            inp = inp.split('_x0')[0] + inp.split('_x0')[1].split('_')[1]
        else:
            inp = inp.split('_X0')[0] + inp.split('_X0')[1].split('_')[1]
    return inp

if __name__ == '__main__':
    # подключаемся к серверу
    cfg = read_config(filename='anketa.ini', section='Mongo')
    conn = MongoClient('mongodb://' + cfg['user'] + ':' + cfg['password'] + '@' + cfg['ip'] + ':' + cfg['port'] + '/'
                       + cfg['db'])
    # выбираем базу данных
    db = conn.saturn_v
    # выбираем коллекцию документов
    colls = db.Products

    # Sort file names with path
    path = "./"
    file_list = os.listdir(path)
    full_list = [os.path.join(path, i) for i in file_list if i.startswith('KEB_') and i.endswith('.xlsx')]
    xlsxs = sorted(full_list, key = os.path.getmtime)

    for xlsx in xlsxs:
        print('\n', xlsx,'\n')
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb[wb.sheetnames[0]]
        wbo = openpyxl.Workbook(write_only=True)
        wso_ish = wbo.create_sheet('Исходный')
        wso_task = wbo.create_sheet('Задание')
        wso_skip_id = wbo.create_sheet('Нет id')
        wso_skip_status = wbo.create_sheet('Нет статуса')
        #wso_double = wbo.create_sheet('Два разных id в одной строке')
        wso_rez = wbo.create_sheet('Результат')
        ids = []
        column = {c: -1 for c in COLUMNS} # генерируем column = {'UTM_CAMPAIGN': -1, ..., 'CARD_STATUS': -1} из COLUMNS
        for i, row in enumerate(ws.values):
            # определяем колонку в которой id
            if not i:
                # заполняем первую строчку вкладок Задание, Нет id, Нет статуса
                wso_task.append(row)
                wso_skip_status.append(row)
                wso_skip_id.append(row)
                #wso_double.append(row)
                for j, cell in enumerate(row):
                    if str(cell).upper() == 'UTM_CAMPAIGN':
                        column['UTM_CAMPAIGN'] = j
                    if str(cell).upper() == 'RESULT':
                        column['RESULT'] = j
                    if str(cell).upper() == 'PRESCORE':
                        column['PRESCORE'] = j
                    if str(cell).upper() == 'LIMIT':
                        column['LIMIT'] = j
                    if str(cell).upper() == 'CARD_STATUS':
                        column['CARD_STATUS'] = j
            else:
                # Если нет нужной информации - выходим
                if column['UTM_CAMPAIGN'] < 0 and (column['RESULT'] < 0 or column['PRESCORE'] < 0):
                    print('Нет столбца UTM_CAMPAIGN, RESULT, LIMIT или PRESCORE')
                    sys.exit()
                # Если не смогли расшифровать remote_id - пропускаем строчку
                remote_id_utm = ''
                if column['UTM_CAMPAIGN'] > -1 and str(type(row[column['UTM_CAMPAIGN']])).find('str') > -1:
                    agent2remote_id = row[column['UTM_CAMPAIGN']]
                    if len(filter_x00(agent2remote_id)[filter_x00(agent2remote_id).find('_') + 1:].strip()) == 36:
                        remote_id_utm = filter_x00(agent2remote_id)[filter_x00(agent2remote_id).find('_') + 1:].strip()
                        if not colls.find({'remote_id': remote_id_utm}).count():
                            wso_skip_id.append(row)
                            row += ('нет такого remote_id в БД:' + str(remote_id_utm),)
                            wso_task.append(row)
                            continue
                if remote_id_utm == '': # Нет id
                    wso_skip_id.append(row)
                    row += ('remote_id не определился: ' + str(row[column['UTM_CAMPAIGN']]),)
                    wso_task.append(row)
                    continue
                # Если не смогли расшифровать статус - пропускаем строчку
                status = -1
                if column['RESULT'] > -1:
                    status = STATUSES['RESULT'].get(filter_x00(row[column['RESULT']]).lower().strip(), -1)
                else:
                    if column['PRESCORE'] > -1:
                        status = STATUSES['PRESCORE'].get(filter_x00(row[column['PRESCORE']]).lower().strip(), -1)
                    if column['LIMIT'] > -1 and (str(filter_x00(row[column['LIMIT']])).find('>') > -1
                                                 or l(filter_x00(row[column['LIMIT']]))):
                        status = STATUSES['LIMIT']['карта выдана']
                    card_status = STATUSES['CARD_STATUS'].get(filter_x00(row[column['CARD_STATUS']]).lower().strip(),
                                                              False)
                    if column['CARD_STATUS'] > -1 and card_status:
                        status = -1
                if status < 0: # Нет статуса
                    wso_skip_status.append(row)
                    continue
                # заполняем вкладку Задание, добавляя туда remote_id
                row += (remote_id_utm,)
                wso_task.append(row)
                # заполняем вкладку Исходный
                for j, coll in enumerate(colls.find({'remote_id': remote_id_utm})):
                    if not j:
                        fields_ish = []
                        for field in coll.keys():
                            if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find(
                                    'int') < 0:
                                fields_ish.append(str(coll.get(field)))
                            else:
                                fields_ish.append(coll.get(field))
                        wso_ish.append(fields_ish)
                # обновляем
                if column['RESULT'] > -1:
                    colls.update({'remote_id': remote_id_utm}, {'$set': {'state_code': status}})
                elif column['LIMIT'] > -1:
                    colls.update({'remote_id': remote_id_utm}, {'$set': {'state_code': status}})
                elif column['PRESCORE'] > -1:
                    colls.update({'remote_id': remote_id_utm}, {'$set': {'state_code': status}})
                else:
                    print('Этой ошибки быть не должно')
                # заполняем вкладку результата
                for j, coll in enumerate(colls.find({'remote_id': remote_id_utm})):
                    if not j:
                        fields_rez = []
                        for field in coll.keys():
                            if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find(
                                    'int') < 0:
                                fields_rez.append(str(coll.get(field)))
                            else:
                                fields_rez.append(coll.get(field))
                        wso_rez.append(fields_rez)
        wbo.save(xlsx.split('KEB_')[0] + 'loaded/' +
                 time.strftime('%Y-%m-%d_%H-%M', time.gmtime(os.path.getmtime(xlsx))) + '_' +
                 xlsx.split('KEB_')[1])
        os.remove(xlsx)


