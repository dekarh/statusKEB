import sys
import time
import os
import openpyxl

from lib import read_config, s, l
from api_statuses import EVA_STATUS
from xlsx2status import filter_x00

def keb(product, colls, path, file=None):
    COLUMNS = ['UTM_CAMPAIGN', 'RESULT', 'PRESCORE', 'LIMIT', 'CARD_STATUS']
    # Проверка наличия колонок (обязательна одна из каждой группы)
    COLUMNS_NEEDS = [['UTM_CAMPAIGN'], ['RESULT', 'PRESCORE', 'LIMIT']]

    STATUSES = {}

    STATUSES['RESULT'] = {
        'отклонено скорингом': EVA_STATUS['STATUS_DENIED'],
        'отклонена банком': EVA_STATUS['STATUS_DENIED'],
        'не соответствует требованиям': EVA_STATUS['STATUS_DENIED'],
        'продукт не нужен': EVA_STATUS['STATUS_DENIED'],
        'отказ клиента': EVA_STATUS['STATUS_DENIED'],
        'регион без представительства КЕБ': EVA_STATUS['STATUS_DENIED'],
        'дубликат (предыдущая заявка)': EVA_STATUS['STATUS_DENIED'],
        'карта выдана': EVA_STATUS['STATUS_ISSUED'],
        'активирована карта': EVA_STATUS['STATUS_ACTIVATED'],
    }

    STATUSES['LIMIT'] = {
        'отклонено скорингом': EVA_STATUS['STATUS_DENIED'],
        'отклонена банком': EVA_STATUS['STATUS_DENIED'],
        'не соответствует требованиям': EVA_STATUS['STATUS_DENIED'],
        'продукт не нужен': EVA_STATUS['STATUS_DENIED'],
        'отказ клиента': EVA_STATUS['STATUS_DENIED'],
        'регион без представительства КЕБ': EVA_STATUS['STATUS_DENIED'],
        'дубликат (предыдущая заявка)': EVA_STATUS['STATUS_DENIED'],
        'карта выдана': EVA_STATUS['STATUS_ISSUED'],
        'активирована карта': EVA_STATUS['STATUS_ACTIVATED'],
    }

    STATUSES['PRESCORE'] = {
        'отказ': EVA_STATUS['STATUS_DENIED'],
        'одобрено': EVA_STATUS['STATUS_APPROVED'],
    }

    # !!! Обрабатывается в последнюю очередь, ДОБАВЛЯТЬ только исключающие обработку статусы
    STATUSES['CARD_STATUS'] = {
        'карта закрыта': True
    }

    if file:
        xlsxs = [file]
    else:
        file_list = os.listdir(path)
        # Sort file names with path
        full_list = [os.path.join(path, i) for i in file_list if i.startswith('KEB_') and i.endswith('.xlsx')]
        xlsxs = sorted(full_list, key=os.path.getmtime)

    for xlsx in xlsxs:
        print('\n', xlsx, '\n')
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb[wb.sheetnames[0]]
        wbo = openpyxl.Workbook(write_only=True)
        wso_ish = wbo.create_sheet('Исходный')
        wso_task = wbo.create_sheet('Задание')
        wso_skip_id = wbo.create_sheet('Нет id')
        wso_skip_status = wbo.create_sheet('Нет статуса')
        # wso_double = wbo.create_sheet('Два разных id в одной строке')
        wso_rez = wbo.create_sheet('Результат')
        ids = []
        column = {c: -1 for c in COLUMNS}  # генерируем column = {'UTM_CAMPAIGN': -1, ..., 'CARD_STATUS': -1} из COLUMNS
        for i, row in enumerate(ws.values):
            # определяем колонку в которой id
            if not i:
                # заполняем первую строчку вкладок Задание, Нет id, Нет статуса
                wso_task.append(row)
                wso_skip_status.append(row)
                wso_skip_id.append(row)
                # wso_double.append(row)
                for column_name in COLUMNS:
                    for j, cell in enumerate(row):
                        if str(cell).upper() == column_name:
                            column[column_name] = j
            else:
                # Если нет нужных столбцов - выходим
                hasnt_columns = ''
                hasnt_all_columns = False
                for columns_group in COLUMNS_NEEDS:
                    hasnt_groups = True
                    hasnt_columns_temp = ''
                    for column_need in columns_group:
                        if column[column_need] < 0 and hasnt_groups:
                            hasnt_groups = True
                            hasnt_columns_temp += column_need + ', '
                        else:
                            hasnt_groups = False
                    if hasnt_groups:
                        hasnt_all_columns = True
                        hasnt_columns = hasnt_columns_temp.strip().strip(',')
                        break
                if hasnt_all_columns:
                    print('В .xlsx файле дожен присутстввать хотя бы один столбец из:', hasnt_columns)
                    sys.exit()
                # Если не смогли расшифровать remote_id - пропускаем строчку
                remote_id_utm = ''
                if column['UTM_CAMPAIGN'] > -1 and str(type(row[column['UTM_CAMPAIGN']])).find('str') > -1:
                    agent2remote_id = row[column['UTM_CAMPAIGN']]
                    if len(filter_x00(agent2remote_id)[filter_x00(agent2remote_id).find('_') + 1:].strip()) == 36:
                        remote_id_utm = filter_x00(agent2remote_id)[filter_x00(agent2remote_id).find('_') + 1:].strip()
                        if not colls.find({'remote_id': remote_id_utm}).count():
                            wso_skip_id.append(row)
                            row += ('нет такого remote_id в БД:' + str(remote_id_utm),)
                            wso_task.append(row)
                            continue
                if remote_id_utm == '':  # Нет id
                    wso_skip_id.append(row)
                    row += ('remote_id не определился: ' + str(row[column['UTM_CAMPAIGN']]),)
                    wso_task.append(row)
                    continue
                # Если не смогли расшифровать статус - пропускаем строчку
                status = -1
                if column['RESULT'] > -1:
                    status = STATUSES['RESULT'].get(filter_x00(row[column['RESULT']]).lower().strip(), -1)
                else:
                    if column['PRESCORE'] > -1:
                        status = STATUSES['PRESCORE'].get(filter_x00(row[column['PRESCORE']]).lower().strip(), -1)
                    if column['LIMIT'] > -1 and (str(filter_x00(row[column['LIMIT']])).find('>') > -1
                                                 or l(filter_x00(row[column['LIMIT']]))):
                        status = STATUSES['LIMIT']['карта выдана']
                    card_status = STATUSES['CARD_STATUS'].get(filter_x00(row[column['CARD_STATUS']]).lower().strip(),
                                                              False)
                    if column['CARD_STATUS'] > -1 and card_status:
                        status = -1
                if status < 0:  # Нет статуса
                    wso_skip_status.append(row)
                    continue
                # заполняем вкладку Задание, добавляя туда remote_id
                row += (remote_id_utm,)
                wso_task.append(row)
                # заполняем вкладку Исходный
                for j, coll in enumerate(colls.find({'remote_id': remote_id_utm})):
                    if not j:
                        fields_ish = []
                        for field in coll.keys():
                            if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find(
                                    'int') < 0:
                                fields_ish.append(str(coll.get(field)))
                            else:
                                fields_ish.append(coll.get(field))
                        wso_ish.append(fields_ish)
                # обновляем
                if column['RESULT'] > -1:
                    colls.update({'remote_id': remote_id_utm}, {'$set': {'state_code': status}})
                elif column['LIMIT'] > -1:
                    colls.update({'remote_id': remote_id_utm}, {'$set': {'state_code': status}})
                elif column['PRESCORE'] > -1:
                    colls.update({'remote_id': remote_id_utm}, {'$set': {'state_code': status}})
                else:
                    print('Этой ошибки быть не должно')
                # заполняем вкладку результата
                for j, coll in enumerate(colls.find({'remote_id': remote_id_utm})):
                    if not j:
                        fields_rez = []
                        for field in coll.keys():
                            if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find(
                                    'int') < 0:
                                fields_rez.append(str(coll.get(field)))
                            else:
                                fields_rez.append(coll.get(field))
                        wso_rez.append(fields_rez)
        if file:
            wbo.save('loaded' + product.upper() + '/' +
                     time.strftime('%Y-%m-%d_%H-%M', time.gmtime(os.path.getmtime(xlsx))) + '_' + xlsx)
        else:
            wbo.save('loaded' + product.upper() + '/' +
                     time.strftime('%Y-%m-%d_%H-%M', time.gmtime(os.path.getmtime(xlsx))) + '_' + xlsx.split('KEB_')[1])
        os.remove(xlsx)

if __name__ == '__main__':
    pass
